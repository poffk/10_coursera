from lxml import etree
import requests
from bs4 import BeautifulSoup
import json
import openpyxl
from os.path import join
from random import sample


QUALITY_OF_OUTPUT_COURSES = 20


def get_courses_list():
    xml = requests.get('https://www.coursera.org/sitemap~www~courses.xml')
    root = etree.fromstring(xml.content)
    links = [link.text for link in root.iter('{*}loc')]
    randomized_links = sample(links, QUALITY_OF_OUTPUT_COURSES)
    return randomized_links


def get_course_info(course_url):
    all_course_data = requests.get(course_url).content
    soup = BeautifulSoup(all_course_data, 'html.parser')
    course_name = get_course_name(soup)
    course_rating = get_course_rating(soup)
    course_language = get_course_language(soup)
    course_duration = get_course_duration(soup)
    course_start_date = get_course_start_date(soup)
    return (course_name, course_language, course_rating, course_duration, course_start_date)


def get_course_name(soup):
    if soup.find('h1', {'class': 'course-name-text display-2-text'}):
        return soup.find('h1', {'class': 'course-name-text display-2-text'}).text
    else:
        return soup.find('h1', {'class': 'course-name-text display-2-text long-title'}).text


def get_course_language(soup):
    return soup.find('div', {'class': 'language-info'}).text


def get_course_start_date(soup):
    if soup.find('script', {'type': 'application/ld+json'}):
        json_content_string = soup.find('script', {'type': 'application/ld+json'}).text
        json_content = json.loads(json_content_string)
        return json_content['hasCourseInstance'][0]['startDate']
    else:
        return None


def get_course_duration(soup):
    return len(soup.find_all('div', {'class': 'week-heading body-2-text'}))


def get_course_rating(soup):
    if soup.find('div', {'class': 'ratings-text bt3-hidden-xs'}):
        return soup.find('div', {'class': 'ratings-text bt3-visible-xs'}).text
    else:
        return None


def output_courses_info_to_xlsx(filepath, courses_info):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Language'
    sheet['C1'] = 'Rating'
    sheet['D1'] = 'Duration, weeks'
    sheet['E1'] = 'Start date'
    for course_number, course_info in enumerate(courses_info):
        sheet.cell(row=2 + course_number, column=1).value = course_info[0]
        sheet.cell(row=2 + course_number, column=2).value = course_info[1]
        sheet.cell(row=2 + course_number, column=3).value = course_info[2]
        sheet.cell(row=2 + course_number, column=4).value = course_info[3]
        sheet.cell(row=2 + course_number, column=5).value = course_info[4]
    wb.save(join(filepath, 'coursera.xlsx'))


if __name__ == '__main__':
    excel_filepath = input('Введите путь до директории: ')
    print('Отлично, работаем дальше!')
    courses_urls_list = get_courses_list()
    courses_info = [get_course_info(course_url) for course_url in courses_urls_list]
    output_courses_info_to_xlsx(excel_filepath, courses_info)
    print('Отлично, дело сделано!')